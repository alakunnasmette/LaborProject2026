import openpyxl

# Load the Excel template
wb = openpyxl.load_workbook('Loopbaan onderzoek 5.0 template.xlsx')

# Check Phase 2.0 sheet structure
sheet_20 = wb["Fase 2.0 | Loopbaanankers"]
print("=== Phase 2.0 | Loopbaanankers ===")
print(f"Dimensions: {sheet_20.dimensions}")
print("\nFirst 15 rows:")
for row in range(1, 16):
    row_data = []
    for col in range(1, 8):  # Check columns A-G
        cell = sheet_20.cell(row=row, column=col)
        val = cell.value
        if val is None:
            val = ""
        row_data.append(str(val)[:20])
    print(f"Row {row}: {row_data}")

# Check Phase 2.1 sheet structure
sheet_21 = wb["Fase 2.1 | Carriere Clusters"]
print("\n=== Phase 2.1 | Carriere Clusters ===")
print(f"Dimensions: {sheet_21.dimensions}")
print("\nFirst 15 rows:")
for row in range(1, 16):
    row_data = []
    for col in range(1, 8):  # Check columns A-G
        cell = sheet_21.cell(row=row, column=col)
        val = cell.value
        if val is None:
            val = ""
        row_data.append(str(val)[:20])
    print(f"Row {row}: {row_data}")
