import openpyxl
import os
import shutil
from datetime import datetime


def write_assessment_answers_to_excel(
    assessment_results: dict,
    excel_path: str = "Loopbaan onderzoek 5.0 template.xlsx",
) -> str | bool:
    """Write Big Five answers into a copy of the Excel template.

    Mapping: Q1→C4, Q2→D5, Q3→E6, Q4→F7, Q5→G8, Q6→C9, ... (cycle C-G every 5).
    Returns the path to the saved file on success, otherwise False.
    """
    try:
        if not os.path.exists(excel_path):
            print(f"Error: Excel file not found at {excel_path}")
            return False

        folder = os.path.dirname(excel_path) or "."
        filled_dir = os.path.join(folder, "results")
        base_name = os.path.splitext(os.path.basename(excel_path))[0]
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        new_name = f"{base_name} - filled {timestamp}.xlsx"
        new_path = os.path.join(filled_dir, new_name)
        shutil.copy2(excel_path, new_path)

        wb = openpyxl.load_workbook(new_path)
        ws = wb.worksheets[1]

        cols = [3, 4, 5, 6, 7]
        for item_num in sorted(assessment_results.keys()):
            val = assessment_results[item_num]
            pos = (item_num - 1) % 5
            cycle = (item_num - 1) // 5
            col = cols[pos]
            row = 4 + cycle * 5 + pos
            ws.cell(row=row, column=col, value=val)

        wb.save(new_path)
        return new_path

    except Exception as e:
        print(f"Error writing to Excel: {e}")
        return False


# maybe won't need this
def write_loopbaan_answers_to_excel(loopbaan_results: dict, excel_path: str = "Loopbaan onderzoek 5.0 template.xlsx") -> str | bool:
    """Copy template into `results` and return the path; caller may write cells."""
    if not os.path.exists(excel_path):
        return False
    folder = os.path.dirname(excel_path) or "."
    filled_dir = os.path.join(folder, "results")
    base_name = os.path.splitext(os.path.basename(excel_path))[0]
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    new_name = f"{base_name} - filled {timestamp}.xlsx"
    new_path = os.path.join(filled_dir, new_name)
    shutil.copy2(excel_path, new_path)
    return new_path


# continue the same Excel file --------------------------------------

PHASE_SHEETS = {
    "1.1": 1,
    "2.0": 2,
    "2.1": 3,
    "2.2": 4,
}


def open_existing_results_excel(excel_path: str, phase: str):
    """
    Open an existing results Excel file and return (workbook, worksheet)
    for the given phase.
    """
    if not excel_path or not os.path.exists(excel_path):
        print("Error: existing Excel file not found.")
        return None, None

    if phase not in PHASE_SHEETS:
        print(f"Error: unknown phase '{phase}'.")
        return None, None

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.worksheets[PHASE_SHEETS[phase]]
    return wb, ws


# Helpers ----------------------------------------------

def iter_block_rows(start_row: int, block_size: int, gap_size: int, repeats: int):
    """Yield row numbers for repeated row blocks."""
    row = start_row
    for _ in range(repeats):
        for i in range(block_size):
            yield row + i
        row += block_size + gap_size


# Phase 2.0 --------------------------


PHASE_2_0_MAPPING = {
    i: {"row": r, "options": ["C", "D", "E", "F", "G"]}
    for i, r in enumerate([
        4, 7, 10, 15, 18, 21, 26, 30, 32, 35,
        37, 39, 41, 43, 45, 48, 50, 53, 56, 60,
        64, 68, 71, 75, 79, 82, 86, 88, 92, 96
    ], start=1)
}

def write_phase_2_0(ws, answers: dict):
    """
    answers: {question_number: column_letter}
    """
    for q_num, col in answers.items():
        if q_num not in PHASE_2_0_MAPPING:
            continue
        row = PHASE_2_0_MAPPING[q_num]["row"]
        ws[f"{col}{row}"] = 1

def write_career_anchors_to_excel(
    answers: dict,
    excel_path: str
) -> str | bool:
            """
            write Phase 2.0 (Career Anchors) answers into an existing results Excel file.

            answers: {question_number: column_letter}
            excel_path: path returned from Phase 1.1
            """

            try:
                wb, ws = open_existing_results_excel(excel_path, phase="2.0")
                if not wb or not ws:
                    return False

                write_phase_2_0(ws, answers)

                wb.save(excel_path)
                return excel_path

            except Exception as e:
                print(f"Error writing Phase 2.0 to Excel: {e}")
                return False



# Phase 2.1 --------------------------------------------


PHASE_2_1_MAPPING = {
    "C": {"start_row": 4, "block_size": 7, "gap_size": 1, "repeats": 16},
    "E": {"start_row": 4, "block_size": 5, "gap_size": 3, "repeats": 16},
    "G": {"start_row": 4, "block_size": 5, "gap_size": 3, "repeats": 16},
}

def write_phase_2_1(ws, answers: dict):
    """
    answers:
    {
        "C": [0/1, 0/1, ...],
        "E": [...],
        "G": [...]
    }
    """
    for col, cfg in PHASE_2_1_MAPPING.items():
        if col not in answers:
            continue

        rows = list(iter_block_rows(**cfg))
        for row, val in zip(rows, answers[col]):
            if val == 1:
                ws[f"{col}{row}"] = 1


# Phase 2.2 --------------------------------------

PHASE_2_2_MAPPING = {
    "C": {"start_row": 2, "block_size": 4, "gap_size": 1, "repeats": 4}
}


def write_phase_2_2(ws, answers: list[int]):
    """
    answers: [1–4, 1–4, ...]
    """
    cfg = PHASE_2_2_MAPPING["C"]
    rows = list(iter_block_rows(**cfg))

    for row, val in zip(rows, answers):
        ws[f"C{row}"] = val