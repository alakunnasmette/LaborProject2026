import openpyxl
from phases.phase20 import CAREER_STATEMENTS
from phases.phase21 import CLUSTERS

# Load Excel
wb = openpyxl.load_workbook('Loopbaan onderzoek 5.0 template.xlsx')
sheet_20 = wb["Fase 2.0 | Loopbaanankers"]
sheet_21 = wb["Fase 2.1 | Carriere Clusters"]

# Check Phase 2.0 - Find where each statement text appears
print("=== PHASE 2.0 MAPPING ===")
print(f"Total CAREER_STATEMENTS: {len(CAREER_STATEMENTS)}")
print(f"Expected row_ids: 1-{len(CAREER_STATEMENTS)}")

# Find statement texts in Excel
statement_positions = {}
for row in range(4, 125):
    cell_b = sheet_20.cell(row=row, column=2).value  # Column B (Stelling)
    if cell_b and isinstance(cell_b, str) and len(str(cell_b)) > 5:
        # This might be a statement
        # Try to match with our statements
        for idx, (num, anchor, text) in enumerate(CAREER_STATEMENTS, start=1):
            if text[:30] in str(cell_b)[:50]:  # Partial match
                statement_positions[idx] = row
                print(f"row_id={idx}: Excel row {row}")
                break

print(f"\nFound {len(statement_positions)} mappings")

# Check Phase 2.1 - Find cluster rows
print("\n=== PHASE 2.1 MAPPING ===")
print(f"Total CLUSTERS: {len(CLUSTERS)}")

# Find where each cluster number appears
cluster_positions = {}
for row in range(4, 132):
    cell_a = sheet_21.cell(row=row, column=1).value  # Column A (Cluster)
    if cell_a and isinstance(cell_a, int):
        cluster_positions[cell_a] = row
        print(f"Cluster {cell_a}: Excel row {row}")

print(f"\nCluster pattern:")
for cid in sorted(cluster_positions.keys())[:3]:
    print(f"Cluster {cid} at row {cluster_positions[cid]}")
    # Check next rows to find the total row
    for check_row in range(cluster_positions[cid], cluster_positions[cid] + 12):
        cell_vals = [sheet_21.cell(row=check_row, column=c).value for c in range(1, 8)]
        print(f"  Row {check_row}: {cell_vals}")
