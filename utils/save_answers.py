import os
from openpyxl import Workbook, load_workbook

def save_assessment_answers(client_id, answers, progress, excel_path):
	"""
	Save answers and progress for a client to an Excel file.
	Args:
		client_id (str): Unique identifier for the client.
		answers (dict): Question-answer pairs.
		progress (int): Current progress (e.g., question number).
		excel_path (str): Path to the Excel file.
	"""
	if os.path.exists(excel_path):
		wb = load_workbook(excel_path)
		ws = wb.active
	else:
		wb = Workbook()
		ws = wb.active
		ws.append(["client_id", "progress", "answers"])

	# Remove previous entry for client_id
	for row in ws.iter_rows(min_row=2):
		if row[0].value == client_id:
			ws.delete_rows(row[0].row)

	ws.append([
		client_id,
		progress,
		str(answers)
	])
	wb.save(excel_path)

def load_assessment_answers(client_id, excel_path):
	"""
	Load saved answers and progress for a client from Excel file.
	Args:
		client_id (str): Unique identifier for the client.
		excel_path (str): Path to the Excel file.
	Returns:
		(progress, answers) or (None, None) if not found.
	"""
	if not os.path.exists(excel_path):
		return None, None
	wb = load_workbook(excel_path)
	ws = wb.active
	for row in ws.iter_rows(min_row=2):
		if row[0].value == client_id:
			progress = row[1].value
			answers = eval(row[2].value) if row[2].value else {}
			return progress, answers
	return None, None
