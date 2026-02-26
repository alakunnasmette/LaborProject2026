import openpyxl
import os
import shutil
from datetime import datetime


# Phase 1.1 --------------------------
def write_assessment_answers_to_excel(
    assessment_results: dict,
    excel_path: str = "Loopbaan onderzoek 5.0 template.xlsx",
) -> str | bool:
    """Write Big Five answers into a copy of the Excel template.

    Mapping: Q1→C4, Q2→D5, Q3→E6, Q4→F7, Q5→G8, Q6→C9, ... (cycle C-G every 5).
    Returns the path to the saved file on success, otherwise False.
    """
    try:
        if not os.path.exists(excel_path):
            print(f"Error: Excel file not found at {excel_path}")
            return False

        folder = os.path.dirname(excel_path) or "."
        filled_dir = os.path.join(folder, "results")
        base_name = os.path.splitext(os.path.basename(excel_path))[0]
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        new_name = f"{base_name} - filled {timestamp}.xlsx"
        new_path = os.path.join(filled_dir, new_name)
        shutil.copy2(excel_path, new_path)

        wb = openpyxl.load_workbook(new_path)
        ws = wb.worksheets[1]

        cols = [3, 4, 5, 6, 7]
        for item_num in sorted(assessment_results.keys()):
            val = assessment_results[item_num]
            pos = (item_num - 1) % 5
            cycle = (item_num - 1) // 5
            col = cols[pos]
            row = 4 + cycle * 5 + pos
            ws.cell(row=row, column=col, value=val)

        wb.save(new_path)
        return new_path

    except Exception as e:
        print(f"Error writing to Excel: {e}")
        return False


# # maybe won't need this
# def write_loopbaan_answers_to_excel(loopbaan_results: dict, excel_path: str = "Loopbaan onderzoek 5.0 template.xlsx") -> str | bool:
#     """Copy template into `results` and return the path; caller may write cells."""
#     if not os.path.exists(excel_path):
#         return False
#     folder = os.path.dirname(excel_path) or "."
#     filled_dir = os.path.join(folder, "results")
#     base_name = os.path.splitext(os.path.basename(excel_path))[0]
#     timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
#     new_name = f"{base_name} - filled {timestamp}.xlsx"
#     new_path = os.path.join(filled_dir, new_name)
#     shutil.copy2(excel_path, new_path)
#     return new_path


# continue the same Excel file --------------------------------------

PHASE_SHEETS = {
    "1.1": 1,
    "2.0": 2,
    "2.1": 3,
    "2.2": 4,
}


def open_existing_results_excel(excel_path: str, phase: str):
    """
    Open an existing results Excel file and return (workbook, worksheet)
    for the given phase.
    """
    if not excel_path or not os.path.exists(excel_path):
        print("Error: existing Excel file not found.")
        return None, None

    if phase not in PHASE_SHEETS:
        print(f"Error: unknown phase '{phase}'.")
        return None, None

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.worksheets[PHASE_SHEETS[phase]]
    return wb, ws


# Helpers ----------------------------------------------

def iter_block_rows(start_row: int, block_size: int, gap_size: int, repeats: int):
    """Yield row numbers for repeated row blocks."""
    row = start_row
    for _ in range(repeats):
        for i in range(block_size):
            yield row + i
        row += block_size + gap_size


# Phase 2.0 --------------------------


PHASE_2_0_MAPPING = {
    i: {"row": r, "options": ["C", "D", "E", "F", "G"]}
    for i, r in enumerate([
        4, 7, 10, 15, 18, 21, 26, 30, 32, 35,
        37, 39, 41, 43, 45, 48, 50, 53, 56, 60,
        64, 68, 71, 75, 79, 82, 86, 88, 92, 96
    ], start=1)
}

def write_phase_2_0(ws, answers: dict):
    """
    answers: {question_number: column_letter}
    """
    for q_num, col in answers.items():
        if q_num not in PHASE_2_0_MAPPING:
            continue
        row = PHASE_2_0_MAPPING[q_num]["row"]
        ws[f"{col}{row}"] = 1

def write_career_anchors_to_excel(
    answers: dict,
    excel_path: str
) -> str | bool:
            """
            write Phase 2.0 (Career Anchors) answers into an existing results Excel file.

            answers: {question_number: column_letter}
            excel_path: path returned from Phase 1.1
            """

            try:
                wb, ws = open_existing_results_excel(excel_path, phase="2.0")
                if not wb or not ws:
                    return False

                write_phase_2_0(ws, answers)

                wb.save(excel_path)
                return excel_path

            except Exception as e:
                print(f"Error writing Phase 2.0 to Excel: {e}")
                return False



# Phase 2.1 --------------------------------

PHASE_2_1_MAPPING = {
    "C": {"start_row": 4, "block_size": 7, "gap_size": 1, "repeats": 16},
    "E": {"start_row": 4, "block_size": 5, "gap_size": 3, "repeats": 16},
    "G": {"start_row": 4, "block_size": 5, "gap_size": 3, "repeats": 16},
}


def write_career_clusters_to_excel(answers: dict, excel_path: str) -> str | bool:
    try:
        wb, ws = open_existing_results_excel(excel_path, phase="2.1")
        if not wb or not ws:
            return False

        cfg_c = PHASE_2_1_MAPPING["C"]
        cfg_e = PHASE_2_1_MAPPING["E"]
        cfg_g = PHASE_2_1_MAPPING["G"]

        print("---- PHASE 2.1 DEBUG ----")
        print("Length:", len(answers))

        for (cluster_id, idx), (main, skill, interest) in answers.items():
        
        

            # ---------- COLUMN C ----------
            row_c = (
                cfg_c["start_row"]
                + (cluster_id - 1) * (cfg_c["block_size"] + cfg_c["gap_size"])
                + (idx - 1)
            )

            ws[f"C{row_c}"] = main


            # ---------- COLUMN E ----------
            if idx <= cfg_e["block_size"]:  # only 5 rows per cluster
                row_e = (
                    cfg_e["start_row"]
                    + (cluster_id - 1) * (cfg_e["block_size"] + cfg_e["gap_size"])
                    + (idx - 1)
                )
                ws[f"E{row_e}"] = skill


            # ---------- COLUMN G ----------
            if idx <= cfg_g["block_size"]:  # only 5 rows per cluster
                row_g = (
                    cfg_g["start_row"]
                    + (cluster_id - 1) * (cfg_g["block_size"] + cfg_g["gap_size"])
                    + (idx - 1)
                )
                ws[f"G{row_g}"] = interest

        wb.save(excel_path)
        return excel_path

    except Exception as e:
        print(f"Error writing Phase 2.1 clusters to Excel: {e}")
        return False


# Alias for backward compatibility / existing calls
def write_phase_2_1_to_excel(answers: dict, excel_path: str) -> str | bool:
    return write_career_clusters_to_excel(answers, excel_path)


# Phase 2.2 --------------------------------------

PHASE_2_2_MAPPING = {
    "C": {"start_row": 2, "block_size": 4, "gap_size": 1, "repeats": 4}
}

def write_phase_2_2_to_excel(answers: dict, excel_path: str) -> str | bool:
    """
    Expects answers as {(group_id, stmt_index): value, ...}
    Writes them into Excel column C according to PHASE_2_2_MAPPING.
    """
    try:
        wb, ws = open_existing_results_excel(excel_path, "2.2")
        if not wb or not ws:
            return False

        cfg = PHASE_2_2_MAPPING["C"]

        for (gid, idx), val in answers.items():
            row = cfg["start_row"] + (gid - 1) * (cfg["block_size"] + cfg["gap_size"]) + (idx - 1)
            ws[f"C{row}"] = int(val)

        wb.save(excel_path)
        return excel_path

    except Exception as e:
        print(f"Error writing Phase 2.2 to Excel: {e}")
        return False

# Phase 2.3 --------------------------------------

def add_job_characteristics_to_excel(excel_file_path: str, jcm_answers: dict) -> bool:
    """Add job characteristics model (JCM) text answers to existing assessment Excel file.
    
    Args:
        excel_file_path: Path to the filled assessment Excel file
        jcm_answers: Dict of {question_num: answer_text}
    
    Returns:
        True on success, False on error.
    """
    try:
        if not os.path.exists(excel_file_path):
            print(f"Error: Excel file not found at {excel_file_path}")
            return False
        
        wb = openpyxl.load_workbook(excel_file_path)
        
        # Find the sheet "Fase 2.3 | J.C.M."
        sheet_name = "Fase 2.3 | J.C.M."
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found in workbook")
            return False
        
        ws = wb[sheet_name]

        # Write answers to the correct rows

        # Unmerge the A11:D11 range if it exists (to allow writing to D11)
        merged_ranges_to_remove = []
        for merged_range in ws.merged_cells.ranges:
            if 'A11:D11' in str(merged_range) or (merged_range.min_row == 11 and merged_range.max_col == 4):
                merged_ranges_to_remove.append(merged_range)
        
        for merged_range in merged_ranges_to_remove:
            ws.unmerge_cells(str(merged_range))

        # Write answers to the correct rows
        # Row structure: Question 1 → D2, Question 2 → D4, Question 3 → D6, etc.
        # Column D (4) is "Toelichting"
        for question_num in sorted(jcm_answers.keys()):
            answer_text = jcm_answers[question_num]
            row = question_num * 2  # Gives 2, 4, 6, 8, 10
            col = 4  # Column D
            ws.cell(row=row, column=col, value=answer_text)
        
        wb.save(excel_file_path)
        return True
        
    except Exception as e:
        print(f"Error adding job characteristics to Excel: {e}")
        return False
