import openpyxl


def read_results_from_excel(filepath):

    wb = openpyxl.load_workbook(filepath, data_only=True)

    user_data = {}

    # ---------------- PHASE 1.1 BIG FIVE ----------------
    ws = wb.worksheets[1]  # same sheet index used in writer

    traits_map = {
        "Extraversie": "C54",
        "Altruisme": "D54",
        "Conscientieusheid": "E54",
        "Neuroticisme": "F54",
        "Openheid": "G54",
    }

    bigfive = {}

    for trait, cell in traits_map.items():
        score = ws[cell].value

        # Simple interpretation (can adjust later if needed)
        if score is None:
            level = "neutral"
        elif score >= 20:
            level = "high"
        elif score <= 10:
            level = "low"
        else:
            level = "neutral"

        bigfive[trait] = {
            "score": score,
            "total": 30,
            "level": level
        }

    user_data["bigfive"] = bigfive


    # ---------------- PHASE 2.0 LOOPBAANANKERS ----------------
    ws = wb.worksheets[2]

    anchors_map = {
        "OMHOOG_KOMEN": ws["C100"].value,
        "VEILIG_VOELEN": ws["D100"].value,
        "VRIJ_ZIJN": ws["E100"].value,
        "BALANS_VINDEN": ws["F100"].value,
        "UITDAGING_ZOEKEN": ws["G100"].value,
    }

    sorted_anchors = sorted(
        anchors_map.items(),
        key=lambda x: x[1] if x[1] is not None else 0,
        reverse=True
    )

    user_data["loopbaanankers"] = [
        sorted_anchors[0][0],
        sorted_anchors[1][0],
    ]


    # ---------------- PHASE 2.1 CARRIERE CLUSTERS ----------------
    ws = wb.worksheets[3]

    clusters_map = {
        "MILIEU": ws["H4"].value,
        "ARCHITECTUUR": ws["H12"].value,
        "KUNST": ws["H20"].value,
        "BUSINESS": ws["H28"].value,
        "EDUCATIE": ws["H36"].value,
        "FINANCIEN": ws["H44"].value,
        "OVERHEID": ws["H52"].value,
        "GEZONDHEIDSWETENSCHAPPEN": ws["H60"].value,
        "HOSPITALITY": ws["H68"].value,
        "HUMANITAIRE": ws["H76"].value,
        "ICT": ws["H84"].value,
        "VEILIGHEID": ws["H92"].value,
        "FABRICAGE": ws["H100"].value,
        "MARKETING": ws["H108"].value,
        "STEM": ws["H116"].value,
        "TRANSPORT": ws["H124"].value,
    }

    sorted_clusters = sorted(
        clusters_map.items(),
        key=lambda x: x[1] if x[1] is not None else 0,
        reverse=True
    )

    user_data["carriereclusters"] = [
        sorted_clusters[0][0],
        sorted_clusters[1][0],
    ]


    # ---------------- PHASE 2.2 CULTUUR ANALYSE ----------------
    ws = wb.worksheets[4]

    cultures_map = {
        "MENSGERICHTE": ws["D2"].value,
        "INNOVATIEVE": ws["D7"].value,
        "BEHEERSGERICHTE": ws["D12"].value,
        "RESULTAATGERICHTE": ws["D17"].value,
    }

    sorted_cultures = sorted(
        cultures_map.items(),
        key=lambda x: x[1] if x[1] is not None else 0,
        reverse=True
    )

    user_data["cultures"] = [
        sorted_cultures[0][0],
        sorted_cultures[1][0],
    ]


    # ---------------- PHASE 2.3 JCM ----------------
    ws = wb.worksheets[5]

    jcm = {
        "Taakvaardigheid": ws["D2"].value,
        "Taakidentiteit": ws["D4"].value,
        "Taakbetekenis": ws["D6"].value,
        "Autonomie": ws["D8"].value,
        "Feedback": ws["D10"].value,
    }

    user_data["jcm"] = jcm


    return user_data