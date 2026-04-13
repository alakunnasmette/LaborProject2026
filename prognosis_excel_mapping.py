"""
prognosis_excel_mapping.py
--------------------------
Stappen:
1. Copy template on submission
2. Map out answers to Excel
3. Write answers from questionnaire into Excel
4. Testing
"""

import shutil
import os
from openpyxl import load_workbook


# ─────────────────────────────────────────────
# INSTELLINGEN 
# ─────────────────────────────────────────────

TEMPLATE_PATH = "Integratie_Prognose_Model_5.0(2).xlsx"  # het originele template
OUTPUT_FOLDER = "clients"                           # map waar klant-excels komen

# ─────────────────────────────────────────────
# MAPPING
# (indeling gemaakt door Mette - niet aanpassen)
# ─────────────────────────────────────────────

PROGNOSIS_MAPPING = {

    "q1": {
        "depends_on": "gender",
        "man":   {"row": 3,   "options": 3},
        "woman": {"row": 20,  "options": 3},
    },

    "q3":        {"row": 37,  "options": 3},
    "q4":        {"row": 39,  "options": 12},
    "q5":        {"row": 41,  "options": 3},
    "q6":        {"row": 45,  "options": 5},
    "q7":        {"row": 48,  "options": 5},
    "q8":        {"row": 52,  "options": 4},
    "q9":        {"row": 68,  "options": 5},
    "q10":       {"row": 70,  "options": 3},
    "q11":       {"row": 74,  "options": 3},
    "q12_ww":    {"row": 76,  "options": 5},
    "q12_pw":    {"row": 78,  "options": 5},
    "q13":       {"row": 80,  "options": 5},
    "q14":       {"row": 82,  "options": 3},
    "q15":       {"row": 84,  "options": 2},
    "q16":       {"row": 87,  "options": 4},
    "q17":       {"row": 90,  "options": 3},
    "q18":       {"row": 93,  "options": 3},
    "q19":       {"row": 95,  "options": 3},
    "q20":       {"row": 97,  "options": 4},
    "q21":       {"row": 100, "options": 3},
    "q22":       {"row": 103, "options": 3},
    "q23":       {"row": 106, "options": 3},
    "q24":       {"row": 109, "options": 3},

    "q25": {
        "depends_on": "gender",
        "man":   {"row": 113, "options": 4},
        "woman": {"row": 116, "options": 4},
    },

    "q26":       {"row": 119, "options": 5},
    "q27":       {"row": 122, "options": 3},
    "q28":       {"row": 125, "options": 6},
    "q29":       {"row": 127, "options": 3},
    "q30":       {"row": 129, "options": 3},
    "q31":       {"row": 131, "options": 3},
    "q32":       {"row": 133, "options": 3},
    "q33":       {"row": 135, "options": 3},
    "q34":       {"row": 137, "options": 3},
    "q35":       {"row": 140, "options": 4},
    "q36":       {"row": 143, "options": 4},
    "q37":       {"row": 146, "options": 2},
    "q38":       {"row": 148, "options": 2},
    "q39":       {"row": 150, "options": 2},
    "q40":       {"row": 152, "options": 4},
    "q41":       {"row": 154, "options": 2},
    "q42":       {"row": 156, "options": 4},
    "q43":       {"row": 158, "options": 3},
    "q44":       {"row": 160, "options": 4},
    "q45":       {"row": 162, "options": 2},
}

# ─────────────────────────────────────────────
# STAP 1: COPY TEMPLATE
# ─────────────────────────────────────────────

def copy_template(output_path: str):
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    shutil.copy2(TEMPLATE_PATH, output_path)

# ─────────────────────────────────────────────
# STAP 2 & 3: MAP ANSWERS + WRITE TO EXCEL
# ─────────────────────────────────────────────

def fill_excel(file_path: str, answers: dict):
    """
    answers = {
        "q1":  0,    <- 0 = Man, 1 = Vrouw
        "q3":  1,    <- index van gekozen optie (0 = eerste keuze)
        "q4":  4,
        ...
    }
    """
    wb = load_workbook(file_path)
    ws = wb["Sheet"]

    gender = "man" if answers.get("q1", 0) == 0 else "woman"

    for q, config in PROGNOSIS_MAPPING.items():
        if q not in answers:
            continue

        if "depends_on" in config:
            config = config[gender]

        row = config["row"]
        col = chr(ord("D") + answers[q])   # index 0 -> D, 1 -> E, 2 -> F ...

        value = ws[f"{col}{row}"].value
        ws[f"Q{row}"] = value

    wb.save(file_path)


# ─────────────────────────────────────────────
# GECOMBINEERDE FUNCTIE - gebruik deze vanuit de app
# ─────────────────────────────────────────────

def process_submission(client_id: str, answers: dict) -> str:
    """
    Dit is de enige functie die de rest van de app hoeft aan te roepen.

    Wat het doet:
        1. Maakt een kopie van het template voor deze klant
        2. Schrijft de antwoorden naar die kopie
        3. Geeft het bestandspad terug zodat de site het kan opslaan/tonen

    Parameters:
        client_id  -- unieke ID van de klant vanuit het klantaccount
                      bijv. "klant_042" of "jan_de_vries"
        answers    -- dict met vraagnummer -> gekozen index
                      { "q1": 0, "q3": 1, "q4": 4, ... }

    Gebruik vanuit de rest van de app:

        from prognosis_excel_mapping import process_submission

        bestand = process_submission(
            client_id = client.id,
            answers   = form.answers
        )
        # -> "clients/klant_042/prognose.xlsx"

    Geeft terug:
        Pad naar het opgeslagen Excel-bestand (str)
    """
    output_path = os.path.join(OUTPUT_FOLDER, client_id, "prognose.xlsx")
    copy_template(output_path)
    fill_excel(output_path, answers)
    return output_path

# ─────────────────────────────────────────────
# STAP 4: TESTING
# ─────────────────────────────────────────────

def test_mapping():
    """
    Simuleert een volledig ingevuld formulier en controleert
    of alle Q-cellen correct worden ingevuld.

    Draai met:  python prognosis_excel_mapping.py
    """

    # Scenario: Man, WW-uitkering, MBO opleiding
    # Conditionele vragen die bewust ontbreken:
    #   q12_pw -> alleen bij bijstand
    #   q23    -> alleen bij HBO
    #   q24    -> alleen bij WO
    #   q28    -> alleen bij naturalisatieproces
    #   q36    -> alleen bij werkloosheidsroute II
    test_answers = {
        "q1":      0,   # Man
        "q3":      1,   # Westers
        "q4":      1,   # Utrecht
        "q5":      0,   # Autochtoon
        "q6":      2,   # Extraversie
        "q7":      3,   # Zakelijk Formeel
        "q8":      0,   # Geen problematiek
        "q9":      4,   # Geen beperkingen
        "q10":     2,   # Actief
        "q11":     0,   # WW
        "q12_ww":  1,   # 3-6 maanden WW
        "q13":     2,   # Geen schulden
        "q14":     2,   # Geen overige inkomsten
        "q15":     0,   # Financiele steun beschikbaar
        "q16":     1,   # Normaal IQ
        "q17":     1,   # B1-B2 Nederlands
        "q18":     1,   # 1 vreemde taal
        "q19":     1,   # B1-B2 vreemde taal
        "q20":     1,   # Basis ICT
        "q21":     1,   # MBO
        "q22":     2,   # Goede MBO-perspectieven
        "q25":     0,   # Alleenstaand
        "q26":     2,   # Midden inkomen
        "q27":     1,   # MBO onderwijsniveau
        "q29":     1,   # Gemiddeld netwerk
        "q30":     2,   # Geen justitieproblemen
        "q31":     2,   # Geen thuisproblemen
        "q32":     2,   # Geen verslaving
        "q33":     2,   # Rijbewijs ja
        "q34":     2,   # Onbeperkte mobiliteit
        "q35":     0,   # Betaald -> Betaald
        "q37":     0,   # Homogene transitie
        "q38":     0,   # Volledig werkloos
        "q39":     0,   # Vrijwilligerswerk
        "q40":     1,   # 3-6 maanden werkloos
        "q41":     0,   # Structureel werk
        "q42":     1,   # 6-10 jaar ervaring
        "q43":     1,   # 5-10 sollicitaties/mnd
        "q44":     0,   # Jan-Mrt
        "q45":     0,   # Man (sector)
    }

    if not os.path.exists(TEMPLATE_PATH):
        print(f"[FOUT] Template niet gevonden: {TEMPLATE_PATH}")
        print(f"       Zorg dat '{TEMPLATE_PATH}' in dezelfde map staat.")
        return

    output = process_submission("test_klant", test_answers)

    wb = load_workbook(output, data_only=True)
    ws = wb["Integratie Prognose Model"]

    verwacht = {
        3: "q1",  37: "q3",  39: "q4",  41: "q5",  45: "q6",
        48: "q7", 52: "q8",  68: "q9",  70: "q10", 74: "q11",
        76: "q12_ww", 80: "q13", 82: "q14", 84: "q15",
        87: "q16", 90: "q17", 93: "q18", 95: "q19", 97: "q20",
        100: "q21", 103: "q22", 113: "q25", 119: "q26",
        122: "q27", 127: "q29", 129: "q30", 131: "q31",
        133: "q32", 135: "q33", 137: "q34", 140: "q35",
        146: "q37", 148: "q38", 150: "q39", 152: "q40",
        154: "q41", 156: "q42", 158: "q43", 160: "q44", 162: "q45",
    }

    conditioneel = ["q12_pw", "q23", "q24", "q28", "q36"]

    print("\n=== TEST RESULTATEN ===")
    fouten = 0
    for row, vraag in verwacht.items():
        val = ws[f"Q{row}"].value
        ok  = val is not None
        if not ok:
            fouten += 1
        print(f"  Q{row:>3} ({vraag:<8}): {'OK' if ok else 'LEEG'}  ->  {val}")

    print(f"\n  Overgeslagen (conditioneel): {', '.join(conditioneel)}")
    print(f"\n  {'GESLAAGD' if fouten == 0 else f'MISLUKT  --  {fouten} onverwacht lege cellen'}")
    print(f"  Opgeslagen als: {output}\n")


# ─────────────────────────────────────────────
# START
# ─────────────────────────────────────────────

if __name__ == "__main__":
    test_mapping()